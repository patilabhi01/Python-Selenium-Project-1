import openpyxl


class HomePageData:

    test_HomePage_data= [{"firstname":"Abhishek","lastname":"Patil","gender":"Male"},{"firstname":"Harika","lastname":"Edla","gender":"Female"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict={}
        book = openpyxl.load_workbook("C:\\Users\\Abhishek Patil\\Documents\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):  # to get coloums
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]